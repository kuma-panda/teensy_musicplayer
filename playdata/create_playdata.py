import sys
import os
import glob
import json
import struct
import collections
import contextlib
import mutagen
from mutagen.mp3 import MP3
from mutagen.mp4 import MP4

import openpyxl as Excel

from pathlib import Path
from PIL import Image

def write_byte(fp, value):
    fp.write(struct.pack('B', value))

def read_byte(fp):
    return struct.unpack('B', fp.read(1))[0]

def write_word(fp, value):
    fp.write(struct.pack('<H', value))

def read_word(fp):
    return struct.unpack('<H', fp.read(2))[0]

def write_string(fp, s):
    s = s.encode('utf-8')
    write_byte(fp, len(s))
    fp.write(s)

def read_string(fp):
    n = read_byte(fp)
    return fp.read(n).decode('utf-8')

def convert_code(s):
    if chr(0xFF5E) in s:
        print('  WARNING: U+FF5E found -- 「{}」'.format(s))
        s = s.replace(chr(0xFF5E), chr(0x301C))
    return s

def color_565(r, g, b):
    return ((r & 0xF8) << 8) | ((g & 0xFC) << 3) | (b >> 3)

class Song:
    def __init__(self, album):
        self.__album = album
        self.__filename = ''
        self.__duration = 0
        self.__title = ''
        self.__track_index = 0
        self.__bitrate = 0
        self.__sample_rate = 0
    
    @property
    def filename(self):
        return self.__filename
    @property
    def duration(self):
        return self.__duration
    @property
    def title(self):
        return self.__title
    @property
    def track_index(self):
        return self.__track_index

    def load(self, path):
        self.__filename = os.path.basename(path)
        ext = os.path.splitext(path)[-1].lower()
        if 'mp3' in ext:
            f = MP3(path)
            self.__duration = round(f.info.length)
            self.__bitrate = int(round(f.info.bitrate / 1000))
            self.__sample_rate = f.info.sample_rate // 100
            self.__album.artist.name = convert_code(str(f['TPE1'].text[0]))
            self.__album.title = convert_code(str(f['TALB'].text[0]))
            self.__title = convert_code(str(f['TIT2'].text[0]))
            self.__album.year = int(str(f['TDRC'].text[0]))
            self.__track_index = int(str(f['TRCK'].text[0]).split('/')[0])
            # print('   {} ({} {})'.format(self.__title, self.__bitrate, self.__sample_rate))
        elif 'm4a' in ext:
            f = MP4(path)
            self.__duration = round(f.info.length)
            self.__bitrate = int(round(f.info.bitrate / 1000))
            self.__sample_rate = f.info.sample_rate // 100
            self.__album.artist.name = convert_code(f['\xa9ART'][0])
            self.__album.title = convert_code(f['\xa9alb'][0])
            self.__title = convert_code(f['\xa9nam'][0])
            self.__album.year = int(f['\xa9day'][0])
            self.__track_index = int(f['trkn'][0][0])
            # print('   {} ({} {})'.format(self.__title, self.__bitrate, self.__sample_rate))
        else:
            raise Exception('Unsupported format -- {}'.format(ext))
    
    def write_binary(self, fp):
        write_string(fp, self.__filename)
        write_word(fp, self.__duration)
        write_word(fp, self.__bitrate)      # kbps 単位 (ex. 320)
        write_word(fp, self.__sample_rate)  # 100Hz 単位 (ex. 441)
        write_string(fp, self.__title)
    
    def get_all_chars(self):
        return list(self.__title)

class Album:
    def __init__(self, artist):
        self.__gid = 0
        self.__artist = artist
        self.__folder_name = ''
        self.__songs = []
        self.__title = ''
        self.__year = 0
        self.__cover_image = None
        self.__thumb_image = None

    @property
    def gid(self):
        return self.__gid
    @gid.setter
    def gid(self, value):
        self.__gid = value
    @property
    def folder_name(self):
        return self.__folder_name
    @property
    def artist(self):
        return self.__artist
    @property
    def title(self):
        return self.__title
    @title.setter
    def title(self, value):
        self.__title = value
    @property
    def year(self):
        return self.__year
    @year.setter
    def year(self, value):
        self.__year = value
    @property
    def songs(self):
        return self.__songs

    def load(self, album_directory_path):
        self.__folder_name = os.path.split(album_directory_path)[-1]
        search_path = os.path.join(album_directory_path, '*')
        song_files = [f for f in glob.glob(search_path) if os.path.splitext(f)[-1].lower() in ['.mp3', '.m4a']]
        for song_path in song_files:
            song = Song(self)
            song.load(song_path)
            self.__songs.append(song)
        if not self.__songs:
            return False
        self.__songs.sort(key=lambda s: s.track_index)

        image_path = os.path.join(album_directory_path, 'coverart.png')
        if os.path.isfile(image_path):
            img = Image.open(image_path)
            cover = img.resize((150, 150), Image.LANCZOS)
            self.__cover_image = cover.convert('RGB')
            thumb = img.resize((60, 60), Image.LANCZOS)
            self.__thumb_image = thumb.convert('RGB')
        else:
            raise Exception('Cover Art Image not found -- {}'.format(album_directory_path))
        self.save_album_data()
        return True

    def write_binary(self, fp):
        write_word(fp, self.__gid)
        write_string(fp, self.__folder_name)
        write_string(fp, self.__title)
        write_word(fp, self.__year)
        write_byte(fp, len(self.__songs))
        write_word(fp, sum([song.duration for song in self.__songs]))
        # ここには「曲数」「総演奏時間」のみを書き込む。各曲の情報は書き込まない

    def save_album_data(self):
        file_path = os.path.join(self.__artist.directory, self.__folder_name, 'album.bin')
        with open(file_path, mode='wb') as fp:
            write_byte(fp, 1 if 'm4a' in self.__songs[0].filename.lower() else 0)
            write_byte(fp, len(self.__songs))
            for song in self.__songs:
                song.write_binary(fp)
            self.write_image(fp)
        print('  album data created -- {}'.format(file_path))

    def write_image(self, fp):
        # image_path = os.path.join(self.__artist.directory, self.__folder_name, 'coverart.bin')
        # with contextlib.suppress(FileNotFoundError):
        #     os.remove(image_path)
        img = self.__cover_image.convert('RGB')
        w, h = img.size
        # write_word(fp, w)
        # write_word(fp, h)
        for y in range(h):
            for x in range(w):
                r, g, b = img.getpixel((x, y))
                write_word(fp, color_565(r, g, b))                     

    def create_thumbnail(self, save_dir):
        img = self.__thumb_image.convert('RGB')
        w, h = img.size
        path = os.path.join(save_dir, '{}.thb'.format(self.__gid))
        with open(path, mode='wb') as fp:
            for y in range(h):
                for x in range(w):
                    r, g, b = img.getpixel((x, y))
                    write_word(fp, color_565(r, g, b))                     

    def get_all_chars(self):
        chars = list(self.__title)
        for song in self.__songs:
            chars += song.get_all_chars()
        return chars

class Artist:
    def __init__(self, artist_directory_path):
        self.__gid = 0
        self.__folder_path = artist_directory_path
        self.__name = ''
        self.__albums = []

    @property
    def gid(self):
        return self.__gid
    @property
    def directory(self):
        return self.__folder_path
    @property
    def folder_name(self):
        return os.path.split(self.__folder_path)[-1]
    @property
    def name(self):
        return self.__name
    @name.setter
    def name(self, value):
        self.__name = value
    @property
    def albums(self):
        return self.__albums

    def load(self):
        search_path = os.path.join(self.__folder_path, '*')
        album_directories = [f for f in glob.glob(search_path) if os.path.isdir(f)]
        for album_dir in album_directories:
            album = Album(self)
            if album.load(album_dir):
                self.__albums.append(album)
        self.__albums.sort(key=lambda a: (a.year, a.folder_name))
        return True if self.__albums else False

    def set_gid(self, value):
        self.__gid = value
        value += 1
        for album in self.__albums:
            album.gid = value
            value += 1
        return value

    def write_binary(self, fp):
        write_word(fp, self.__gid)
        write_string(fp, self.folder_name)
        write_string(fp, self.__name)
        write_byte(fp, len(self.__albums))
        for album in self.__albums:
            album.write_binary(fp)

    def create_thumbnail(self, save_dir):
        image_path = os.path.join(self.__folder_path, 'artist.png')
        if not os.path.isfile(image_path):
            raise Exception('Cover Art Image not found -- {}'.format(album_directory_path))
        img = Image.open(image_path)
        img = img.resize((60, 60), Image.LANCZOS)
        img = img.convert('RGB')
        thumb_path = os.path.join(save_dir, '{}.thb'.format(self.__gid))
        with open(thumb_path, mode='wb') as fp:
            for y in range(60):
                for x in range(60):
                    r, g, b = img.getpixel((x, y))
                    write_word(fp, color_565(r, g, b))                     
        for album in self.__albums:
            album.create_thumbnail(save_dir)

    def get_all_chars(self):
        chars = list(self.__name)
        for album in self.__albums:
            chars += album.get_all_chars()
        return chars

if __name__ == '__main__':
    root_directory = sys.argv[1]
    search_path = os.path.join(root_directory, '*')
    artist_directories = [dir for dir in glob.glob(search_path) if os.path.isdir(dir)]
    artists = []
    chars = [chr(c) for c in range(0x20, 0x7F)]
    gid = 0
    for dir in artist_directories:
        print('processing: {}'.format(dir))
        artist = Artist(dir)
        if artist.load():
            gid = artist.set_gid(gid)
            artists.append(artist)
    artists.sort(key=lambda a: a.folder_name.lower())
    binary_path = os.path.join(root_directory, 'playdata.bin')
    with open(binary_path, mode='wb') as fp:
        write_byte(fp, len(artists))
        for artist in artists:
            artist.write_binary(fp)
            chars += artist.get_all_chars()
        print('{} successfully created.'.format(binary_path))
    # charlist_path = os.path.join(root_directory, 'charlist.txt')
    # with open(charlist_path, mode='w', encoding='utf-8') as fp:
    #     chars = list(collections.Counter(chars).keys())
    #     chars.sort()
    #     fp.write(''.join(chars))
    #     print('{} characters extracted to {}'.format(len(chars), charlist_path))

    # サムネイル画像を保存（場所は root_directory/thumbs）
    thumbnail_dir = os.path.join(root_directory, 'thumbs')
    for artist in artists:
        artist.create_thumbnail(thumbnail_dir)

    # リードバックして確認
    book = Excel.Workbook()
    sheet = book.active
    row = 1
    for artist in artists:
        sheet.cell(column=1, row=row).value = artist.gid
        sheet.cell(column=2, row=row).value = artist.name
        sheet.cell(column=3, row=row).value = artist.folder_name
        for album in artist.albums:
            sheet.cell(column=4, row=row).value = album.gid
            sheet.cell(column=5, row=row).value = album.title
            sheet.cell(column=6, row=row).value = album.folder_name
            sheet.cell(column=7, row=row).value = album.year
            for song in album.songs:
                sheet.cell(column=8, row=row).value = song.track_index
                sheet.cell(column=9, row=row).value = song.title
                sheet.cell(column=10, row=row).value = song.filename
                sheet.cell(column=11, row=row).value = song.duration
                row += 1
    book.save('playdata.xlsx')
